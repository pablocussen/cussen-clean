"""
Script para extraer APUs de archivos Excel de ONDA y Proyecto Loica
y crear una base de datos consolidada en JSON
"""

import json
import os
from pathlib import Path

import openpyxl


def extract_apu_from_excel(file_path):
    """
    Extrae información de APU de un archivo Excel
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        apu_data = {
            "filename": os.path.basename(file_path),
            "category": get_category_from_path(file_path),
            "items": [],
            "total": None,
            "unit": None,
        }

        # Intentar extraer datos de las primeras 50 filas
        for row in range(1, min(51, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(11, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                value = cell.value
                if value is not None:
                    row_data.append(str(value))

            if row_data:
                apu_data["items"].append(row_data)

        return apu_data

    except Exception as e:
        print(f"Error procesando {file_path}: {str(e)}")
        return None


def get_category_from_path(file_path):
    """
    Extrae la categoría del path del archivo
    """
    parts = Path(file_path).parts

    # Buscar categorías conocidas
    categories = [
        "HORMIGONES",
        "ALBANILERIA",
        "REVESTIMIENTOS",
        "MOLDAJES",
        "ENFIERRADURAS",
        "MOVIMIENTO TIERRA",
        "FAENA",
        "PAVIMENTOS",
    ]

    for part in parts:
        part_upper = part.upper()
        for cat in categories:
            if cat in part_upper:
                return cat

    return "VARIOS"


def scan_apu_folder(base_path):
    """
    Escanea la carpeta APU y extrae todos los archivos Excel
    """
    apu_database = {"onda_apus": [], "loica_project": None, "categories": {}}

    excel_files = list(Path(base_path).rglob("*.xlsx"))

    print(f"📊 Encontrados {len(excel_files)} archivos Excel")

    for i, excel_file in enumerate(excel_files, 1):
        print(f"Procesando {i}/{len(excel_files)}: {excel_file.name}")

        apu_data = extract_apu_from_excel(str(excel_file))

        if apu_data:
            # Determinar si es Loica o ONDA
            if "Loica" in excel_file.name or "Presupuesto" in excel_file.name:
                apu_database["loica_project"] = apu_data
            else:
                apu_database["onda_apus"].append(apu_data)

                # Organizar por categoría
                category = apu_data["category"]
                if category not in apu_database["categories"]:
                    apu_database["categories"][category] = []
                apu_database["categories"][category].append(apu_data["filename"])

    return apu_database


def create_simplified_database(apu_database):
    """
    Crea una base de datos simplificada con las actividades más comunes
    """
    simplified_db = {
        "actividades": [
            {
                "id": "radier_10cm",
                "nombre": "Radier e=10cm, 212.5 kg/m3",
                "unidad": "m2",
                "categoria": "HORMIGONES",
                "materiales": [
                    {"nombre": "Cemento", "cantidad": 5.0, "unidad": "sacos"},
                    {"nombre": "Arena", "cantidad": 0.55, "unidad": "m3"},
                    {"nombre": "Ripio", "cantidad": 0.85, "unidad": "m3"},
                    {"nombre": "Malla Acma C-188", "cantidad": 1.1, "unidad": "m2"},
                ],
                "mano_obra": [
                    {"nombre": "Jornal", "cantidad": 0.3, "unidad": "HH"},
                    {"nombre": "Ayudante", "cantidad": 0.3, "unidad": "HH"},
                ],
                "rendimiento": "3.5 m2/HH",
                "precio_referencia": 18500,
            },
            {
                "id": "muro_fiscal_15cm",
                "nombre": "Albañilería Ladrillo Fiscal e=15cm",
                "unidad": "m2",
                "categoria": "ALBANILERIA",
                "materiales": [
                    {"nombre": "Ladrillo Fiscal", "cantidad": 65, "unidad": "un"},
                    {"nombre": "Mortero pega", "cantidad": 0.042, "unidad": "m3"},
                    {"nombre": "Cemento", "cantidad": 10.5, "unidad": "kg"},
                    {"nombre": "Arena", "cantidad": 0.042, "unidad": "m3"},
                ],
                "mano_obra": [
                    {"nombre": "Albañil", "cantidad": 1.2, "unidad": "HH"},
                    {"nombre": "Ayudante", "cantidad": 0.6, "unidad": "HH"},
                ],
                "rendimiento": "0.8 m2/HH",
                "precio_referencia": 24500,
            },
            {
                "id": "excavacion_zanja",
                "nombre": "Excavación zanja a brazo, ancho ≤1.4m, prof ≤2m",
                "unidad": "m3",
                "categoria": "MOVIMIENTO TIERRA",
                "materiales": [],
                "mano_obra": [{"nombre": "Jornal", "cantidad": 2.5, "unidad": "HH"}],
                "rendimiento": "0.4 m3/HH",
                "precio_referencia": 12000,
            },
            {
                "id": "moldaje_muro",
                "nombre": "Moldaje muro (2.5 usos)",
                "unidad": "m2",
                "categoria": "MOLDAJES",
                "materiales": [
                    {"nombre": "Madera pino bruto", "cantidad": 0.024, "unidad": "m3"},
                    {"nombre": 'Clavos 3"', "cantidad": 0.25, "unidad": "kg"},
                    {"nombre": "Separadores", "cantidad": 4, "unidad": "un"},
                ],
                "mano_obra": [
                    {"nombre": "Carpintero", "cantidad": 0.8, "unidad": "HH"},
                    {"nombre": "Ayudante", "cantidad": 0.8, "unidad": "HH"},
                ],
                "rendimiento": "1.25 m2/HH",
                "precio_referencia": 8500,
            },
            {
                "id": "enfierradura_d10",
                "nombre": "Enfierradura D=10mm A44-28",
                "unidad": "kg",
                "categoria": "ENFIERRADURAS",
                "materiales": [
                    {"nombre": "Fierro D=10mm", "cantidad": 1.05, "unidad": "kg"},
                    {"nombre": "Alambre negro", "cantidad": 0.015, "unidad": "kg"},
                ],
                "mano_obra": [{"nombre": "Fierrero", "cantidad": 0.08, "unidad": "HH"}],
                "rendimiento": "12.5 kg/HH",
                "precio_referencia": 1850,
            },
            {
                "id": "estuco_exterior",
                "nombre": "Estuco exterior sobre albañilería",
                "unidad": "m2",
                "categoria": "REVESTIMIENTOS",
                "materiales": [
                    {"nombre": "Cemento", "cantidad": 8, "unidad": "kg"},
                    {"nombre": "Arena", "cantidad": 0.035, "unidad": "m3"},
                    {"nombre": "Hidrorrepelente", "cantidad": 0.05, "unidad": "lt"},
                ],
                "mano_obra": [
                    {"nombre": "Estucador", "cantidad": 0.6, "unidad": "HH"},
                    {"nombre": "Ayudante", "cantidad": 0.3, "unidad": "HH"},
                ],
                "rendimiento": "1.7 m2/HH",
                "precio_referencia": 7500,
            },
        ],
        "metadata": {
            "fuente": "APU ONDA 2019 + Proyecto Loica 2024",
            "total_apus_onda": len(apu_database.get("onda_apus", [])),
            "categorias": list(apu_database.get("categories", {}).keys()),
            "fecha_actualizacion": "2024",
        },
    }

    return simplified_db


if __name__ == "__main__":
    import io
    import sys

    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

    print("EXTRACTOR DE APUs - CLAUDIA SODIMAC")
    print("=" * 60)

    base_path = r"c:\Users\pablo\claudia_bot\APU"

    # Escanear carpeta APU
    print("\nEscaneando carpeta APU...")
    apu_database = scan_apu_folder(base_path)

    # Crear base de datos simplificada
    print("\nCreando base de datos simplificada...")
    simplified_db = create_simplified_database(apu_database)

    # Guardar resultados
    output_path = r"c:\Users\pablo\claudia_bot\web_app\apu_database.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(simplified_db, f, ensure_ascii=False, indent=2)

    print(f"\nBase de datos creada: {output_path}")
    print(f"Total APUs ONDA procesados: {len(apu_database.get('onda_apus', []))}")
    print(f"Categorias encontradas: {', '.join(apu_database.get('categories', {}).keys())}")
    print(f"Actividades en DB simplificada: {len(simplified_db['actividades'])}")

    # Guardar también el escaneo completo
    full_output_path = r"c:\Users\pablo\claudia_bot\APU\apu_scan_full.json"
    with open(full_output_path, "w", encoding="utf-8") as f:
        json.dump(apu_database, f, ensure_ascii=False, indent=2)

    print(f"Escaneo completo guardado: {full_output_path}")
    print("\nProceso completado!")
